
import re
from time import sleep
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

import openpyxl
import time

from modulos import fazer_login
from modulos import config

opts = webdriver.ChromeOptions()
opts.add_argument("--disable-extensions")
# opts.add_argument("--headless")

s = Service('./chromedriver.exe')

opts.add_experimental_option('excludeSwitches', ['enable-logging']) #desabilitar mensagem de DevTools listening on ws

driver = webdriver.Chrome(service=s, options=opts)

if not fazer_login.run(driver):
    print('Usuário ou senha incorretos')
    driver.quit()
    exit()

# # driver.get(config.BASE_URL + '/sipac/consultarHistorico.do?acao=200')


# Lendo dados da planilha
workbook = openpyxl.load_workbook("tombamentos.xlsx")
sheet = workbook.active

erros = []

for row in sheet.iter_rows(min_row=2, max_col=12, values_only=True):
    if row[0] is not None:
        driver.get(config.BASE_URL + '/sipac/tombaBemAntigo.do?acao=100&aba=gerencia-menupatrimonio')
        el_caption = driver.find_element(By.XPATH, '//*[@id="formulario"]/table[1]/caption')
        print(f'Cadastrando tombamento {row[0]}')
        el_tombamento = driver.find_element(By.XPATH, '//*[@id="numTombamento"]') # Nº de tombamento
        el_tombamento.send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
        el_tombamento.send_keys(str(row[0])) 

        el_material = driver.find_element(By.XPATH, '//*[@id="model"]') # Material
        el_material.send_keys(row[1]) 
        sleep(1)
        el_material.send_keys(Keys.TAB)
        sleep(0.3)
        el_material.send_keys(Keys.TAB)
        sleep(1)

        el_marca = driver.find_element(By.XPATH, '//*[@id="nome"]') # Marca
        el_marca.send_keys(row[2])
        sleep(1)
        el_marca.send_keys(Keys.TAB)

        sleep(1)
        driver.find_element(By.XPATH, '//*[@id="valor"]').send_keys(str("%.2f" % float(row[3]))) # Valor estimado
        sleep(0.3)
        driver.find_element(By.XPATH, '//*[@id="estado"]').send_keys(row[4]) # Estado do bem 
        sleep(0.3)
        driver.find_element(By.XPATH, '//*[@id="finalidade"]').send_keys(row[5]) # Finalidade
        sleep(0.3)
        el_unid_resp = driver.find_element(By.XPATH, '//*[@id="codigo_unid_id"]') # Unidade responsável
        el_unid_resp.send_keys(row[6]) 
        el_unid_resp.send_keys(Keys.TAB)
        sleep(0.3)
        el_num_termo = driver.find_element(By.XPATH, '//*[@id="numTermo"]')
        el_num_termo.send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
        el_num_termo.send_keys(row[7]) # Num termo  
        sleep(0.3)
        el_ano_termo = driver.find_element(By.XPATH, '//*[@id="anoTermo"]')
        el_ano_termo.send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
        el_ano_termo.send_keys(row[8]) # Ano termo  
        sleep(0.3)
        driver.find_element(By.NAME, 'bem.observacoes').send_keys(row[9]) # Observações
        sleep(0.3)

        driver.find_element(By.XPATH, '//*[@id="formulario"]/table[2]/tfoot/tr/td/center/input[1]').click()
        sleep(1)
        driver.switch_to.alert.accept()

        el_erros = driver.find_elements(By.XPATH, '//*[@color="red"]')
        if(len(el_erros)>0):
            for erro in el_erros:
                erros.append({'tombo': row, 'erro': erro.text})
    
        
workbook.close()

if len(erros) > 0:
    print(f"\n Foram encontrados {len(erros)} erros ao tentar realizar o cadastro \n")

    wb = openpyxl.Workbook()

    new_wb = wb.active
    new_wb.append(['Nº Tombamento','Material','Marca','Valor','Estimado','Situação','Finalidade','Unidade','Responsável','TR / 2023','Ano do TR','Observação', 'erro']
)
    for itens, erro in enumerate(erros):
        new_wb.append(erros[itens]['tombo']+(erros[itens]['erro'],))
        
    try:
        file_name = f"\\erros-tombamento.xlsx"
        wb.save(config.work_dir + file_name)
        print(f" > Planilha com erros gravada em {file_name}")
    except:
        print(f"ERRO ao tentar salvar a planilha com erros {file_name}")

    wb.close()

print(f"\n>>>> Fim de execução do script\n")
driver.quit()
exit()


tombamentos = {}

for row in sheet.iter_rows(min_row=2,  max_col=1):
    for celula in row:
        if celula.value is not None:
            search_box = driver.find_element(By.XPATH, '//*[@id="numTombamento"]')
            print(f'Buscando {celula.value}')
            search_box.send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
            sleep(0.3)
            search_box.send_keys(celula.value)
            sleep(0.3)
            driver.find_element(By.XPATH, '//*[@id="formulario"]/table[2]/tfoot/tr/td/input[1]').click()
            sleep(1)
            historico = driver.find_element(By.XPATH, '//*[@id="conteudo"]/table[2]/tbody/tr[2]/td[1]/a').get_attribute('onclick')
            regex_historico = r'\/sipac\/consultarHistorico\.do\?popup=true&numero=\d+&tipoConsulta=\d+&acao=\d+'
            termoguia = driver.find_element(By.XPATH, '//*[@id="conteudo"]/table[2]/tbody/tr[2]/td[3]/a').get_attribute('onclick')
            regex_termoguia = r'\/sipac\/consultarTermoGuia\.do\?tipoTombamentoTermo=\d+&popup=true&tipoConsulta=\d+&numero=\d+&ano=\d+'
            tombamentos[celula.value] = {'tombo': celula.value,
                                        'historico': str(re.search(regex_historico, historico).group(0)),
                                       'termoguia': re.search(regex_termoguia, termoguia).group(0),
                                       'unidade': driver.find_element(By.XPATH, '//*[@id="conteudo"]/table[2]/tbody/tr[1]/td').text,
                                       'denominacao': driver.find_element(By.XPATH, '//*[@id="conteudo"]/table[2]/tbody/tr[2]/td[2]').text,
                                       'termo': driver.find_element(By.XPATH, '//*[@id="conteudo"]/table[2]/tbody/tr[2]/td[3]').text,
                                       'valor': driver.find_element(By.XPATH, '//*[@id="conteudo"]/table[2]/tbody/tr[2]/td[4]').text
                                       }

workbook.close()

for tombo, itens in tombamentos.items():

    print(f'Acessando a página de histórico do tombo {tombo}')
    driver.get(config.BASE_URL + tombamentos[tombo]['historico'])  
    el_usuario = driver.find_element(By.XPATH, '//*[@id="container-popup"]/table[1]/tbody/tr[26]/td')             
    tombamentos[tombo]['usuario'] = el_usuario.text
    
    print(f'Acessando a página de Termo de responsabilidade do tombo {tombo}')
    driver.get(config.BASE_URL + tombamentos[tombo]['termoguia'])    
    el_fornecedor = driver.find_element(By.XPATH, '//*[@id="container-popup"]/table[1]/tbody/tr[5]/td/table/tbody/tr[1]/td')
    tombamentos[tombo]['fornecedor'] = el_fornecedor.text

driver.quit()

print("\nGravando planilha com informações do levantamento")


